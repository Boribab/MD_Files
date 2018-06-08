'''
사람이 엑셀에서
- 데이터가 들어있는 파일 실행
- 데이터가 들어있는 시트로 이동
- 데이터가 있는 위치의 데이터를 활용

프로그램이 엑셀에서
- 데이터가 들어있는 파일명으로 클래스 변수 생성
- 클래스 변수에서 시트 이름을 활용하여 시트 이동
- 데이터가 있는 위치의 데이터를 활용
'''
from openpyxl import load_workbook

# load_workbook 함수를 이용하여 엑셀 클래스 변수 생성
wb = load_workbook('test1.xlsx')
# 활성화 된 시트를 sheet 변수로 설정 
sheet = wb.active

print(sheet['A1'].value)
print(sheet['A2'].value)
print(sheet['B1'].value)
print(sheet['C1'].value)
print(sheet['D3'].value)