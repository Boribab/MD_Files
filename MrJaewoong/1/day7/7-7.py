from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
sheet.title = 'My_class'
# 시트 이름을 title을 써서 바꿀 수 있다.

sheet.append(['Number', 'name'])

for i in range(10):
	sheet.append([i, chr(i+65)*3])

wb.save('test3.xlsx')

# ======================================
# 저장 된 엑셀 파일의 내용을 출력해보자

rows = sheet['1:11']
for row in rows:
	print(row[0].value, row[1].value)

# 행 번호를 일일이 확인할 수 없으므로
# 셀이 입력되어 잇는 구간을 알아서 인식하도록 하면 좋음
# 이럴 때 iter_rows()를 쓰면 간단히 해결

for row in sheet.iter_rows():
	print(row[0].value, row[1].value)